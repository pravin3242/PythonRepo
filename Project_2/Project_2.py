from datetime import datetime
import pandas as pd
from playwright.sync_api import sync_playwright

# Read configuration from Excel file
df = pd.read_excel("C:\\Users\\lnv0179\\Desktop\\Python\\Input123.xlsx",
                   sheet_name='Configuration')
username = df.loc[0, 'Username']
password = df.loc[0, 'Password']
keywd = df.loc[0, 'Keyword']

# Define column names for DataFrame
column_names = ["Job Summary", "Description", "Skills", "About Client", "URL"]

# Create an empty DataFrame to store the scraped data
result_df = pd.DataFrame(columns=column_names)


def scrape_job_data(page):
    # Extract job data
    href = page.locator("//div//a[@data-test='slider-open-in-new-window UpLink']").get_attribute('href')
    URL = "https://www.upwork.com" + href
    title = page.locator("//h4[@class='d-flex align-items-center mt-0 mb-5']").text_content()
    description = page.locator("//div[@data-test='Description']").text_content()
    skills = page.locator("//section[@data-test='Expertise']").text_content()
    about_client = page.locator(
        "//div[@data-test='about-client-container AboutClientUserShared AboutClientUser']").text_content()

    # Append data to DataFrame
    result_df.loc[len(result_df)] = [title, description, skills, about_client, URL]
    print(result_df)


def run(playwright):
    # Launch browser and create new context and page
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()

    # Navigate to login page
    page.goto("https://www.upwork.com/ab/account-security/login")

    # Fill in login details and log in
    page.get_by_placeholder("Username or Email").click()
    page.get_by_placeholder("Username or Email").fill(username)
    page.get_by_role("button", name="Continue", exact=True).click()
    page.get_by_role("textbox", name="Password").fill(password)
    page.get_by_role("button", name="Log in").click()

    # Close dialogs and open search
    page.get_by_role("dialog").get_by_role("button", name="Close", exact=True).click()
    page.get_by_label("Close").click()
    page.get_by_placeholder("Search for jobs").click()
    page.get_by_placeholder("Search for jobs").fill(keywd)
    page.get_by_placeholder("Search for jobs").press("Enter")

    # Wait for dropdown to load and select option
    page.wait_for_timeout(3000)
    page.click("//div[@data-test='jobs_per_page UpCDropdown']")
    page.wait_for_timeout(3000)
    page.keyboard.press('ArrowDown')
    page.keyboard.press('ArrowDown')
    page.keyboard.press('Enter')
    page.wait_for_timeout(5000)

    # Extract job data
    elements = page.locator(
        "//article//div//small//span[contains(text(),'minutes ago')] | //article//div//small//span[contains(text(),'hours ago')] | //article//div//small//span[contains(text(),'hour ago')]").element_handles()

    # Iterate through job data elements, scrape data, and close popups
    for element in elements:
        element.click()
        scrape_job_data(page)
        page.keyboard.press('Escape')  # Close the job detail popup

    # Close browser context and browser
    context.close()
    browser.close()


# Run the script using Playwright
with sync_playwright() as playwright:
    run(playwright)

# Generate a sheet name based on current date and time
current_time = datetime.now()
sheet_name = current_time.strftime('Sheet_%Y-%m-%d_%H-%M-%S')

# Print the scraped data
print(result_df)

# Write DataFrame to Excel file
# Path to your Excel file
file_path = "C:\\Users\\lnv0179\\Desktop\\Python\\Input123.xlsx"

# Attempt to write to a new sheet in the existing workbook
try:
    # Attempt to write to a new sheet in the existing workbook
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
        result_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # If the file does not exist, create it with the DataFrame
except FileNotFoundError:
    # If the file does not exist, create it with the DataFrame
    result_df.to_excel(file_path, sheet_name=sheet_name, index=False)

################################################################################################
# Count of Keyword values

file_path = "C:\\Users\\lnv0179\\Desktop\\Python\\Input123.xlsx"
df = pd.read_excel(file_path, sheet_name='ScoreConfig')

print(df)

score_config_dict = df.set_index('ColumnName').apply(lambda x: {'Keyword': x['Keyword'], 'ScoreValue': x['ScoreValue']},
                                                     axis=1).to_dict()

# Display the dictionary
print(score_config_dict)

# Initialize a dictionary to hold the total count of keywords for each section
total_keyword_counts = {}

# Iterate over each section in the dictionary
for section, data in score_config_dict.items():
    # Split the 'Keyword' string to get individual keywords and count them
    keyword_count = len(data['Keyword'].split(','))
    # Assign the count to the corresponding section
    total_keyword_counts[section] = keyword_count
    print("Test")
    print(section)
    print(data)
    print(total_keyword_counts)

# Output the total keyword counts for each section
for section, count in total_keyword_counts.items():
    print(f"{section}: {count}")

##############################################################################################
# Occurrence of keywords in the list


import pandas as pd
from openpyxl.reader.excel import load_workbook

# Read the Excel file
file_path = "C:\\Users\\lnv0179\\Desktop\\Python\\Input123.xlsx"
df = pd.read_excel(file_path, sheet_name='Result')

# Define keywords for each column
keywords = {
    "Job Summary": "job".lower(),
    "Description": "description".lower(),
    "skills": "skills"
}

# Iterate through rows and calculate total count for each row
total_counts = []
for index, row in df.iterrows():
    row_total_count = 0
    print(row)
    for column, keyword in keywords.items():
        row_total_count += str(row[column]).lower().count(keyword)
    total_counts.append(row_total_count)

# Add a new column "Score" to the DataFrame with the total counts
df['Score'] = total_counts

# Print the DataFrame with the new "Score" column
print(df)

with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='Result', index=False)

##############################################################################################
# Colour format according to score point

import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

workbook_Path = "C:\\Users\\lnv0179\\Desktop\\Python\\Input123.xlsx"
worksheet_name = "Result"

# Load the Excel file
workbook = load_workbook(workbook_Path)

# Access a specific worksheet by name
worksheet = workbook[worksheet_name]

column_name = "Score"
# Find the column index based on column name
column_index = None
for cell in worksheet[1]:
    if cell.value == column_name:
        column_index = cell.column
        break
    # If column found, apply color based on condition to entire row
if column_index:
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=column_index)

        # Apply color based on conditions using conditional formatting
        if cell.value is not None:
            if cell.value < 5:
                fill_color = 'FF0000'  # Red color
            elif 5 <= cell.value <= 7:
                fill_color = 'FFA500'  # Orange color
            elif 8 <= cell.value <= 10:
                fill_color = '00FF00'  # Green color

            if fill_color:
                rule = openpyxl.formatting.rule.CellIsRule(operator='equal', formula=[f'A{row}:F{row}'],
                                                           fill=openpyxl.styles.fills.PatternFill(
                                                               start_color=fill_color, end_color=fill_color))
                # Apply conditional formatting to columns A and B for the current row
                worksheet.conditional_formatting.add(f'A{row}:F{row}', rule)

if column_index:
    column_letter = get_column_letter(column_index)

    # Apply auto filter and sort
    worksheet.auto_filter.ref = f"C1:{column_letter}{worksheet.max_row}"
    worksheet.auto_filter.add_sort_condition(f"{column_letter}2:{column_letter}{worksheet.max_row}", descending=True)

    workbook.save(workbook_Path)
