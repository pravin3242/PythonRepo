import openpyxl
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

workbook_Path =  "/home/hp-66/Desktop/Project2/Input123.xlsx"
worksheet_name = "Result"

# Load the Excel file
workbook = load_workbook(workbook_Path)

# Access a specific worksheet by name
worksheet = workbook[worksheet_name]

column_name = "Score"
# Find the column index based on column name
column_index = None
for cell in worksheet[1]:
    if cell.value == column_name:
        column_index = cell.column
        break
    # If column found, apply color based on condition to entire row
if column_index:
    for row in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row, column=column_index)


        # Apply color based on conditions using conditional formatting
        if cell.value is not None:
            if cell.value < 5:
                fill_color = 'FF0000'  # Red color
            elif 5 <= cell.value <= 7:
                fill_color = 'FFA500'  # Orange color
            elif 8 <= cell.value <= 10:
                fill_color = '00FF00'  # Green color

            if fill_color:

                rule = openpyxl.formatting.rule.CellIsRule(operator='equal', formula=[f'A{row}:C{row}'], fill=openpyxl.styles.fills.PatternFill(start_color=fill_color, end_color=fill_color))
                # Apply conditional formatting to columns A and B for the current row
                worksheet.conditional_formatting.add(f'A{row}:C{row}', rule)

if column_index:
    column_letter = get_column_letter(column_index)

    # Apply auto filter and sort
    worksheet.auto_filter.ref = f"C1:{column_letter}{worksheet.max_row}"
    worksheet.auto_filter.add_sort_condition(f"{column_letter}2:{column_letter}{worksheet.max_row}", descending=True)


    workbook.save(workbook_Path)